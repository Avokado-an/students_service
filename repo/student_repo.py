import os
from configparser import ConfigParser
from datetime import datetime
from idlelib.configdialog import is_int

import openpyxl
from openpyxl import Workbook

from entity.specialty import Specialty
from entity.student import Student
from validator.student_validator import is_valid_specialty, is_valid_mark

config_file = 'config.ini'
config_section_name = "file-config"
config = ConfigParser()
config.read(config_file)

students_path = config[config_section_name]['excel-file-name']
sheet_name = config[config_section_name]["sheet-name"]
filters_path = config[config_section_name]["path-to-filter-results"]


def load_spreadsheet_sheet(file_path, spread_sheet_name):
    spreadsheet_file = load_spreadsheet_file(file_path)
    return spreadsheet_file[spread_sheet_name]


def load_spreadsheet_file(file_path):
    try:
        return openpyxl.load_workbook(file_path)
    except IOError:
        print("No spreadsheet provided")


def load_initial_students():
    student_sheet = load_spreadsheet_sheet(students_path, sheet_name)
    students = []
    for student_row in range(2, student_sheet.max_row + 1):
        student = fill_student_properties_from_spreadsheet(student_sheet, student_row)
        students.append(student)
    return students


def add_student(created_student):
    wb = load_spreadsheet_file(students_path)
    ws = wb[sheet_name]
    created_student.student_id = ws.cell(ws.max_row, 1).value + 1
    student_dict = {
        1: created_student.student_id,
        2: created_student.name,
        3: created_student.surname,
        4: created_student.specialty,
        5: created_student.average_mark
    }
    ws.append(student_dict)
    wb.save(students_path)


def delete_student(student_id):
    wb = load_spreadsheet_file(students_path)
    ws = wb[sheet_name]
    for student_row in range(2, ws.max_row + 1):
        if ws.cell(student_row, 1).value == student_id:
            ws.delete_rows(student_row)
            wb.save(students_path)


def create_student_filtered_file(minimal_mark, subject, students):
    date = datetime.today().strftime('%Y-%m-%d')
    create_path_if_not_exist()
    filename = f"filter_for_{subject}_with_mark_above_{str(minimal_mark)}_for_{str(date)}.xlsx"
    wb = Workbook()
    ws = wb.active
    student_id = 1
    header = {
        1: "id",
        2: "name",
        3: "surname"
    }
    ws.append(header)
    for student in students:
        student_dict = {
            1: student.student_id,
            2: student.name,
            3: student.surname
        }
        ws.append(student_dict)
        student_id += 1
    wb.save(filters_path + filename)


def create_path_if_not_exist():
    if not os.path.exists(filters_path):
        os.makedirs(filters_path)


def fill_student_properties_from_spreadsheet(sheet, row):
    student_id = sheet.cell(row, 1).value
    student_name = sheet.cell(row, 2).value
    student_surname = sheet.cell(row, 3).value
    specialty = sheet.cell(row, 4).value
    average_mark = sheet.cell(row, 5).value

    is_valid_student_data = is_int(student_id) and is_valid_specialty(specialty) and is_valid_mark(average_mark)

    if not is_valid_student_data:
        raise ValueError(f"Invalid config file. File failed for id {student_id}")

    student = Student(student_id,
                      student_name,
                      student_surname,
                      getattr(Specialty, specialty),
                      average_mark
                      )
    return student
